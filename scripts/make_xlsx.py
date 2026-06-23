from __future__ import annotations

from pathlib import Path

import openpyxl


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    csv_path = root / "services_pricelist.csv"
    xlsx_path = root / "services_pricelist.xlsx"

    if not csv_path.exists():
        raise FileNotFoundError(csv_path)

    # Simple CSV->XLSX conversion using Excel-friendly defaults
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List"

    for i, line in enumerate(csv_path.read_text(encoding="utf-8").splitlines(), start=1):
        parts = [p.strip() for p in line.split(",")]
        for j, val in enumerate(parts, start=1):
            ws.cell(row=i, column=j, value=val)

    # Make header bold
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    ws.freeze_panes = "A2"
    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")


if __name__ == "__main__":
    main()

